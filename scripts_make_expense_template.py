from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path


def _default_template_path() -> Path:
    base_dir = Path(__file__).resolve().parent
    candidates = [
        base_dir.parent / "세출예산명세_파싱템플릿.xlsx",
        Path.home() / "Desktop" / "2026 예산" / "세출예산명세_파싱템플릿.xlsx",
        Path.home() / "Documents" / "2026 예산" / "세출예산명세_파싱템플릿.xlsx",
    ]
    for candidate in candidates:
        if candidate.parent.exists():
            return candidate
    return base_dir / "세출예산명세_파싱템플릿.xlsx"


def main():
    path = _default_template_path()
    wb = Workbook()
    ws = wb.active
    ws.title = "세출템플릿"

    headers = [
        "사업명",
        "목코드",
        "목명(선택)",
        "세목코드",
        "세목명(선택)",
        "현재예산(원,자동)",
        "기정예산(원,선택)",
        "요약라벨",
        "내역명",
        "기초금액(원)",
        "수량1",
        "단위1",
        "수량2",
        "단위2",
        "수량3",
        "단위3",
    ]
    ws.append(headers)

    sample_rows = [
        ["충남콘텐츠진흥원 운영", "101", "", "101-01", "", None, "", "소계(총 21명)", "일반직급여(21명)", 1157839000, 1, "년", "", "", "", ""],
        ["충남콘텐츠진흥원 운영", "101", "", "101-02", "", None, "", "소계(총 1명)", "임원급여 (원장, 1명)", 96208000, 1, "년", "", "", "", ""],
        ["충남콘텐츠진흥원 운영", "101", "", "101-03", "", None, "", "소계(총 5명)", "무기직급여(5명)", 268717000, 1, "년", "", "", "", ""],
        ["충남콘텐츠진흥원 운영", "101", "", "101-04", "", None, "", "소계(총 1명)", "기간제급여(1명)", 37734000, 1, "년", "", "", "", ""],
        ["충남콘텐츠진흥원 운영", "107", "", "", "", None, 87216000, "합계", "퇴직급여충당금(퇴직연금)", 95178000, 1, "년", "", "", "", ""],
        ["충남콘텐츠진흥원 운영", "109", "", "", "", None, 50000000, "합계", "경영평가에 따른 성과급(평가급)/직원", 50000000, 1, "년", "", "", "", ""],
    ]
    for row in sample_rows:
        ws.append(row)

    max_rows = 300
    for r in range(2, max_rows + 1):
        ws[f"F{r}"] = f'=IF(J{r}="","",J{r}*IF(K{r}="",1,K{r})*IF(M{r}="",1,M{r})*IF(O{r}="",1,O{r}))'

    dv = DataValidation(type="list", formula1='"명,회,월,년,식,건,개"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{max_rows}")
    dv.add(f"N2:N{max_rows}")
    dv.add(f"P2:P{max_rows}")

    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="E8F2FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 28,
        "B": 10,
        "C": 16,
        "D": 12,
        "E": 20,
        "F": 18,
        "G": 16,
        "H": 18,
        "I": 34,
        "J": 16,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 10,
        "P": 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=2, max_row=max_rows, min_col=6, max_col=11):
        for cell in row:
            cell.number_format = "#,##0"

    guide = wb.create_sheet("작성가이드")
    guide.append(["입력 규칙"])
    guide.append(["1) 목코드/세목코드만 입력해도 파서가 목명/세목명을 자동 보정합니다."])
    guide.append(["2) 현재예산(원)은 기초금액*수량1*수량2*수량3 자동계산입니다(F열)."])
    guide.append(["3) 수량2/수량3은 비우면 1로 계산됩니다."])
    guide.append(["4) 기정예산은 비워도 됩니다. 필요시만 입력하세요."])
    guide.append(["5) 요약라벨 예: 합계(총 28명), 소계(총 21명)"])
    guide.append(["6) 단위는 L/N/P 열 드롭다운에서 선택하세요."])

    wb.save(path)
    print(str(path))


if __name__ == "__main__":
    main()
