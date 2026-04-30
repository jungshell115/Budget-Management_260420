import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


MAX_ROWS = 1200
INPUT_FILL = PatternFill("solid", fgColor="EAF3FF")


def _workspace_roots(base_dir: Path | None = None) -> list[Path]:
    roots: list[Path] = []
    if base_dir is not None:
        roots.extend([base_dir.parent, base_dir])
    home = Path.home()
    roots.extend([home / "Desktop" / "2026 예산", home / "Documents" / "2026 예산"])

    unique_roots: list[Path] = []
    for root in roots:
        if root not in unique_roots:
            unique_roots.append(root)
    return unique_roots


def _default_output_dir(base_dir: Path) -> Path:
    for root in _workspace_roots(base_dir):
        if root.exists():
            return root
    return base_dir.parent


def _safe_name(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(text or "").strip())


def _to_int(value) -> int:
    s = str(value or "").replace(",", "").strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return int(float(m.group(0))) if m else 0


def _first_existing_path(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists():
            return p
    return None


def _read_csv_fallback(path: Path) -> list[dict[str, str]]:
    for enc in ("utf-8-sig", "cp949"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    return []


def _latest_output_dir(base_dir: Path) -> Path | None:
    out = base_dir / "output"
    if not out.exists():
        return None
    dirs = [p for p in out.iterdir() if p.is_dir()]
    if not dirs:
        return None
    return max(dirs, key=lambda p: p.stat().st_mtime)


def _load_seed_rows_from_existing_template(
    selected_businesses: list[str],
    new_businesses: list[str],
) -> list[dict[str, str | int]]:
    candidates: list[Path] = []
    for root in _workspace_roots():
        candidates.extend(
            [
                root / "세출예산명세_파싱템플릿.xlsx",
                root / "세출예산명세_파싱템플릿.csv",
                root / "시스템" / "2026" / "2026_본예산_세출예산명세_파싱템플릿.xlsx",
                root / "시스템" / "2026" / "2026_본예산_세출예산명세_파싱템플릿.csv",
            ]
        )
    src = _first_existing_path(candidates)
    if src is None:
        return []

    selected_set = {x.strip() for x in selected_businesses if x.strip()}
    out: list[dict[str, str | int]] = []
    ext = src.suffix.lower()
    if ext == ".xlsx":
        try:
            wb = load_workbook(src, data_only=True)
            ws = wb.active
            values = list(ws.values)
            if not values:
                return []
            headers = [str(v or "").strip() for v in values[0]]
            for row_vals in values[1:]:
                row = {headers[i]: ("" if i >= len(row_vals) or row_vals[i] is None else str(row_vals[i])) for i in range(len(headers))}
                biz = str(row.get("사업명", "") or "").strip()
                if not biz:
                    continue
                # 선택한 사업만 반영 (선택 없으면 기존사업 미반영)
                if biz not in selected_set:
                    continue
                mok = str(row.get("목코드", "") or "").strip()
                semok = str(row.get("세목코드", "") or "").strip()
                if not mok and not semok:
                    continue
                out.append(
                    {
                        "사업명": biz,
                        "목코드": mok,
                        "세목코드": semok,
                        "기정예산": (
                            _to_int(row.get("기정예산(원,선택)", row.get("기정예산(원)", "0")))
                            or _calc_from_detail_fields(row)
                        ),
                        "내역명": str(row.get("내역명", "") or "").strip(),
                    }
                )
        except Exception:
            return []
    else:
        rows = _read_csv_fallback(src)
        for row in rows:
            biz = str(row.get("사업명", "") or "").strip()
            if not biz:
                continue
            # 선택한 사업만 반영 (선택 없으면 기존사업 미반영)
            if biz not in selected_set:
                continue
            mok = str(row.get("목코드", "") or "").strip()
            semok = str(row.get("세목코드", "") or "").strip()
            if not mok and not semok:
                continue
            out.append(
                {
                    "사업명": biz,
                    "목코드": mok,
                    "세목코드": semok,
                    "기정예산": (
                        _to_int(row.get("기정예산(원,선택)", row.get("기정예산(원)", "0")))
                        or _calc_from_detail_fields(row)
                    ),
                    "내역명": str(row.get("내역명", "") or "").strip(),
                }
            )
    for b in new_businesses:
        name = b.strip()
        if name:
            out.append({"사업명": name, "목코드": "", "세목코드": "", "기정예산": 0, "내역명": ""})
    return out


def _extract_code_fields(row: dict[str, str]) -> tuple[str, str]:
    semok = str(row.get("세목코드", "") or "").strip()
    mok = str(row.get("목코드", "") or "").strip()
    if semok or mok:
        return mok, semok
    subject = str(row.get("과목", "") or "").strip()
    m_semok = re.match(r"^\s*(\d{3}-\d{2})\b", subject)
    if m_semok:
        return "", m_semok.group(1)
    m_mok = re.match(r"^\s*(\d{3})\b", subject)
    if m_mok:
        return m_mok.group(1), ""
    return "", ""


def _pick_first(row: dict[str, str], keys: list[str]) -> str:
    for k in keys:
        v = str(row.get(k, "") or "").strip()
        if v:
            return v
    return ""


def _calc_from_detail_fields(row: dict[str, str]) -> int:
    base = _to_int(_pick_first(row, ["기초금액(원)", "기초금액"]))
    if base <= 0:
        return 0
    q1 = _to_int(_pick_first(row, ["수량1"]))
    q2 = _to_int(_pick_first(row, ["수량2"]))
    q3 = _to_int(_pick_first(row, ["수량3"]))
    mul = (q1 if q1 > 0 else 1) * (q2 if q2 > 0 else 1) * (q3 if q3 > 0 else 1)
    return base * mul


def _pick_prior_budget(row: dict[str, str]) -> int:
    prior = _to_int(_pick_first(row, ["기정예산(원)", "기정예산", "기정"]))
    if prior > 0:
        return prior
    # 본예산 기준 라인은 기정이 0일 수 있어 현재예산을 기준값으로 사용
    current = _to_int(_pick_first(row, ["현재예산(원)", "현재예산"]))
    if current > 0:
        return current
    # 인코딩/헤더 변형 상황 대응: 행의 숫자 컬럼에서 첫 번째 양수 값을 기준으로 사용
    for v in row.values():
        n = _to_int(v)
        if n > 0:
            return n
    return 0


def load_latest_expense_seed_rows(
    base_dir: Path,
    round_label: str,
    selected_businesses: list[str],
    new_businesses: list[str],
) -> list[dict[str, str | int]]:
    latest = _latest_output_dir(base_dir)
    if latest is None:
        return [{"사업명": b, "목코드": "", "세목코드": "", "기정예산": 0, "내역명": ""} for b in new_businesses]

    # 본예산 세출 CSV를 기준으로 기정예산 lookup 구성
    base_rows = _read_csv_fallback(latest / "세출예산명세서_사업단위_다중단위.csv")
    prior_lookup: dict[tuple[str, str, str], int] = {}
    prior_by_semok: dict[tuple[str, str], int] = {}
    prior_by_mok: dict[tuple[str, str], int] = {}
    for row in base_rows:
        biz = _pick_first(row, ["사업명", "사업", "사업명칭"])
        mok, semok = _extract_code_fields(row)
        if not biz:
            continue
        prior = _pick_prior_budget(row)
        prior_lookup[(biz, mok, semok)] = prior
        if semok:
            prior_by_semok[(biz, semok)] = prior
        if mok:
            prior_by_mok[(biz, mok)] = prior

    seed_from_template = _load_seed_rows_from_existing_template(selected_businesses, new_businesses)
    if seed_from_template:
        # 기존 본예산 템플릿의 기정예산이 비어 있으면 최신 output 세출 CSV 값으로 보강
        for item in seed_from_template:
            if _to_int(item.get("기정예산", 0)) > 0:
                continue
            biz = str(item.get("사업명", "") or "").strip()
            mok = str(item.get("목코드", "") or "").strip()
            semok = str(item.get("세목코드", "") or "").strip()
            item["기정예산"] = (
                prior_lookup.get((biz, mok, semok))
                or (prior_by_semok.get((biz, semok)) if semok else 0)
                or (prior_by_mok.get((biz, mok)) if mok else 0)
                or 0
            )
        return seed_from_template

    csv_candidates: list[Path]
    # 세출 템플릿은 항상 '본예산 작성 내역'을 기본 라인으로 가져온다.
    # (추경에서도 동일 라인 기반으로 변경분만 작성)
    csv_candidates = [latest / "세출예산명세서_사업단위_다중단위.csv"]
    src = _first_existing_path(csv_candidates)
    if src is None:
        return [{"사업명": b, "목코드": "", "세목코드": "", "기정예산": 0, "내역명": ""} for b in new_businesses]

    rows = _read_csv_fallback(src)
    selected_set = {x.strip() for x in selected_businesses if x.strip()}
    out: list[dict[str, str | int]] = []
    for row in rows:
        biz = _pick_first(row, ["사업명", "사업", "사업명칭"])
        if not biz:
            continue
        # 선택한 사업만 반영 (선택 없으면 기존사업 미반영)
        if biz not in selected_set:
            continue
        mok, semok = _extract_code_fields(row)
        if not mok and not semok:
            continue
        prior = _pick_prior_budget(row)
        detail = _pick_first(row, ["내역명", "세부사업", "과목"])
        out.append({"사업명": biz, "목코드": mok, "세목코드": semok, "기정예산": prior, "내역명": detail})

    for b in new_businesses:
        name = b.strip()
        if not name:
            continue
        out.append({"사업명": name, "목코드": "", "세목코드": "", "기정예산": 0, "내역명": ""})
    return out


def _style_header(ws):
    header_fill = PatternFill("solid", fgColor="E8F2FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _clear_snapshot_values_keep_layout(ws):
    """
    사용자 커스텀(v3) 레이아웃/수식은 유지하고, 입력값만 비워 템플릿 상태로 만든다.
    - A~C(구분/사업명/부서) 라벨은 유지
    - 수식 셀은 유지
    - 숫자/텍스트 상수 입력 셀은 제거
    """
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            # 헤더/항목명(1~3행)은 그대로 유지
            if r <= 3:
                continue
            if c <= 3:
                continue
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue
            if val is None:
                continue
            cell.value = None


def _find_master_template(out_dir: Path) -> Path | None:
    candidates = sorted(
        [p for p in out_dir.glob("*_v3.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _make_snapshot_sheet(ws, title: str):
    ws.title = title
    ws.append(["구분", "편성목", "예산액", "기정예산", "증감", "산출기초"])
    ws.append(["(제)충남콘텐츠진흥원", "", 0, 0, 0, ""])
    ws.append(["", "기관운영", 0, 0, 0, ""])
    ws.append(["", "목적사업", 0, 0, 0, ""])
    ws.freeze_panes = "A2"
    _style_header(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 44
    for row in ws.iter_rows(min_row=2, max_row=MAX_ROWS, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"


def _make_supp_snapshot_sheet(ws, round_label: str):
    ws.title = f"{round_label}_기초"
    headers = [
        "구분",
        "사업명",
        "부서",
        "재원유형(자동)",
        "시군비_상세(기정,입력)",
        f"시군비_상세({round_label},입력)",
        "기정예산_합계",
        "기정예산_국비",
        "기정예산_도비",
        "기정예산_시군비",
        "기정예산_자체",
        f"{round_label}_합계",
        f"{round_label}_국비",
        f"{round_label}_도비",
        f"{round_label}_시군비",
        f"{round_label}_자체",
        "증감_합계(자동)",
        "증감_국비(자동)",
        "증감_도비(자동)",
        "증감_시군비(자동)",
        "증감_자체(자동)",
        "비고",
    ]
    ws.append(headers)
    ws.append(
        [
            "(제)충남콘텐츠진흥원",
            "충남메타버스지원센터 운영",
            "사업총괄실",
            None,
            "천안 300,000,000 / 아산 442,000,000",
            "천안 300,000,000 / 아산 409,000,000",
            11056000000,
            4950000000,
            5723000000,
            742000000,
            50000000,
            11023000000,
            4950000000,
            5723000000,
            709000000,
            50000000,
            None,
            None,
            None,
            None,
            None,
            "예시",
        ]
    )

    for r in range(2, MAX_ROWS + 1):
        ws[f"D{r}"] = (
            f'=IF(OR(G{r}<>"",H{r}<>"",I{r}<>"",J{r}<>"",K{r}<>""),'
            f'IF(IFERROR(H{r},0)>0,"보조금",IF(AND(IFERROR(K{r},0)>0,IFERROR(H{r},0)=0,IFERROR(I{r},0)=0,IFERROR(J{r},0)=0),"자체","출연금")),"")'
        )
        ws[f"N{r}"] = f'=IF(L{r}="","",L{r}-IF(G{r}="",0,G{r}))'
        ws[f"O{r}"] = f'=IF(M{r}="","",M{r}-IF(H{r}="",0,H{r}))'
        ws[f"P{r}"] = f'=IF(N{r}="","",N{r}-IF(I{r}="",0,I{r}))'
        ws[f"Q{r}"] = f'=IF(O{r}="","",O{r}-IF(J{r}="",0,J{r}))'
        ws[f"R{r}"] = f'=IF(P{r}="","",P{r}-IF(K{r}="",0,K{r}))'

    ws.freeze_panes = "A2"
    _style_header(ws)
    widths = {
        "A": 18,
        "B": 28,
        "C": 14,
        "D": 14,
        "E": 32,
        "F": 32,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
        "R": 16,
        "S": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=MAX_ROWS, min_col=7, max_col=18):
        for cell in row:
            cell.number_format = "#,##0"
    # 자동 계산 셀만 색상 표시
    for col in ("D", "N", "O", "P", "Q", "R"):
        for r in range(2, MAX_ROWS + 1):
            ws[f"{col}{r}"].fill = INPUT_FILL


def _fill_expense_template_sheet(ws, seed_rows: list[dict[str, str | int]]):
    # 2행부터 초기화
    for r in range(2, MAX_ROWS + 1):
        for c in range(1, 17):
            ws.cell(r, c).value = None

    if seed_rows:
        for idx, s in enumerate(seed_rows[: MAX_ROWS - 1], start=2):
            ws.cell(idx, 1, s.get("사업명", ""))
            ws.cell(idx, 2, s.get("목코드", ""))
            ws.cell(idx, 3, s.get("세목코드", ""))
            ws.cell(idx, 5, s.get("기정예산", 0))
            ws.cell(idx, 8, s.get("내역명", ""))
    else:
        # 선택/신규사업이 없으면 빈 템플릿 유지
        pass

    for r in range(2, MAX_ROWS + 1):
        ws[f"D{r}"] = f'=IF(I{r}="","",I{r}*IF(J{r}="",1,J{r})*IF(L{r}="",1,L{r})*IF(N{r}="",1,N{r}))'
        ws[f"F{r}"] = f'=IF(D{r}="","",D{r}-IF(E{r}="",0,E{r}))'
        ws[f"P{r}"] = f'=IF(ABS(F{r})>0,"반영","미반영")'

def _make_expense_template_sheet(wb: Workbook, seed_rows: list[dict[str, str | int]]):
    if "세출예산명세서_파싱템플릿" in wb.sheetnames:
        ws = wb["세출예산명세서_파싱템플릿"]
        _fill_expense_template_sheet(ws, seed_rows)
        return
    ws = wb.create_sheet("세출예산명세서_파싱템플릿")
    ws.append(
        [
            "사업명(필수)",
            "목코드(선택)",
            "세목코드(선택)",
            "현재예산(원,자동)",
            "기정예산(원,선택)",
            "증감(원,자동)",
            "요약라벨(선택)",
            "내역명(필수)",
            "기초금액(원,필수)",
            "수량1(선택)",
            "단위1(선택)",
            "수량2(선택)",
            "단위2(선택)",
            "수량3(선택)",
            "단위3(선택)",
            "추경반영여부(자동)",
        ]
    )
    _fill_expense_template_sheet(ws, seed_rows)
    dv_unit = DataValidation(type="list", formula1='"명,회,월,년,식,건,개,팀"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add(f"K2:K{MAX_ROWS}")
    dv_unit.add(f"M2:M{MAX_ROWS}")
    dv_unit.add(f"O2:O{MAX_ROWS}")
    ws.freeze_panes = "A2"
    _style_header(ws)


def _make_validation_sheet(wb: Workbook):
    if "자동검증" in wb.sheetnames:
        ws = wb["자동검증"]
        ws.delete_rows(1, ws.max_row or 1)
    else:
        ws = wb.create_sheet("자동검증")
    ws.append(["행번호", "오류유형", "조치가이드", "점검값"])

    # 엑셀 복구 팝업 방지를 위해 수식을 단순화한다.
    for r in range(2, MAX_ROWS + 1):
        ws[f"A{r}"] = f'=IF(\'세출예산명세서_파싱템플릿\'!A{r}="","",ROW(\'세출예산명세서_파싱템플릿\'!A{r}))'
        ws[f"B{r}"] = f'=IF(A{r}="","",IF(\'세출예산명세서_파싱템플릿\'!A{r}="","사업명 누락",IF(AND(\'세출예산명세서_파싱템플릿\'!B{r}="",\'세출예산명세서_파싱템플릿\'!C{r}=""),"목/세목코드 누락",IF(\'세출예산명세서_파싱템플릿\'!H{r}="","내역명 누락",""))))'
        ws[f"C{r}"] = f'=IF(B{r}="","",IF(B{r}="사업명 누락","사업명을 입력하세요",IF(B{r}="목/세목코드 누락","목코드 또는 세목코드를 입력하세요",IF(B{r}="내역명 누락","내역명을 입력하세요","확인 필요"))))'
        ws[f"D{r}"] = f'=IF(B{r}="","",\'세출예산명세서_파싱템플릿\'!A{r}&" / "&\'세출예산명세서_파싱템플릿\'!B{r}&" / "&\'세출예산명세서_파싱템플릿\'!C{r})'

    ws.freeze_panes = "A2"
    _style_header(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38


def _make_guide_sheet(wb: Workbook, year: str, round_label: str, dept: str):
    ws = wb.create_sheet("작성가이드")
    ws.append(["통합 작성/검증 가이드"])
    ws.append([f"- 파일명 규칙: {year}_{round_label}_{dept}.xlsx"])
    ws.append([f"- 본 템플릿은 {round_label} 기준으로 작성되었습니다."])
    ws.append(["- 세출 입력은 '세출예산명세서_파싱템플릿' 탭에서 작성하세요."])
    ws.append(["- 자동검증 탭에서 행번호/오류유형/조치가이드를 먼저 확인하세요."])
    ws.append(["- 추경 반영 사업은 증감(원)이 0이 아닌 행으로 자동 판단합니다."])
    ws.append([f"- 본예산/추경 기초 탭은 현재 회차 기준으로 '{round_label}_기초'를 사용합니다."])
    ws.append(["- 기정예산은 최신 output의 세출 CSV를 기준으로 자동 채움됩니다."])
    ws.column_dimensions["A"].width = 90


def main():
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="본예산/1차추경 통합 입력검증 템플릿 생성")
    parser.add_argument("--year", default="2026", help="연도 (예: 2026)")
    parser.add_argument("--round", default="본예산", help="회차 (본예산 또는 1차추경)")
    parser.add_argument("--dept", default="통합", help="부서명 (예: 사업총괄실)")
    parser.add_argument(
        "--businesses",
        default="",
        help="대상 사업명(쉼표구분). 비우면 최신 세출CSV 전체 대상",
    )
    parser.add_argument(
        "--new-businesses",
        default="",
        help="신규 사업명(쉼표구분). 템플릿에 기정 0으로 추가",
    )
    parser.add_argument(
        "--output-dir",
        default=str(_default_output_dir(base_dir)),
        help="출력 폴더 경로",
    )
    args = parser.parse_args()

    year = _safe_name(args.year)
    round_label = _safe_name(args.round)
    dept = _safe_name(args.dept)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{year}_{round_label}_{dept}.xlsx"
    selected_businesses = [x.strip() for x in str(args.businesses or "").split(",") if x.strip()]
    new_businesses = [x.strip() for x in str(args.new_businesses or "").split(",") if x.strip()]
    seed_rows = load_latest_expense_seed_rows(base_dir, round_label, selected_businesses, new_businesses)

    wb: Workbook
    if round_label != "본예산":
        master = _find_master_template(out_dir)
        if master is not None:
            wb = load_workbook(master)
            if wb.sheetnames:
                first = wb[wb.sheetnames[0]]
                first.title = f"{round_label}_기초"
                _clear_snapshot_values_keep_layout(first)
        else:
            wb = Workbook()
            primary_ws = wb.active
            _make_supp_snapshot_sheet(primary_ws, round_label)
    else:
        wb = Workbook()
        primary_ws = wb.active
        _make_snapshot_sheet(primary_ws, "본예산_기초")

    _make_expense_template_sheet(wb, seed_rows)
    _make_validation_sheet(wb)
    if "작성가이드" in wb.sheetnames:
        wb.remove(wb["작성가이드"])
    _make_guide_sheet(wb, year, round_label, dept)

    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
